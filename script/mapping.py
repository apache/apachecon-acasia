# 将ApacheConAsia讲师信息收集表.xlsx、ApacheConSessions.xlsx、drawing1.xml文件放到项目根目录
# 将会在根目录生成mapping.csv文件

from xml.dom.minidom import parse
from pypinyin import lazy_pinyin
import xml.dom.minidom
import openpyxl
import csv

 
 # 首先定义一个CSV文件
head =["collect_row","pic","zh_name","en_name","position","track","title","mail","sessions_row"]
with open("./mapping.csv","a+",encoding="utf-8",newline="") as f:
   csvf=csv.writer(f)
   csvf.writerow(head)

# 使用minidom解析器打开 XML 文档
DOMTree = xml.dom.minidom.parse("./drawing1.xml")

#打开execl文件
collect=openpyxl.load_workbook("./ApacheConAsia讲师信息收集表.xlsx")
session=openpyxl.load_workbook("./ApacheConSessions.xlsx")

# 在集合中获取数据
collection = DOMTree.documentElement
twoCellAnchors = collection.getElementsByTagName("xdr"+":"+"twoCellAnchor")

# 取两个固定符号之间的字符串
def get_str_btw(str,begin,end):
   par=str.partition(begin)
   return (par[2].partition(end))[0][:]

# 获取row和pic
for twoCellAnchor in twoCellAnchors:
 
   # 获取到row
   from1 = twoCellAnchor.getElementsByTagName('xdr'+':'+'to')[0]
   row1=from1.getElementsByTagName('xdr'+':'+'row')[0]
   row=int(row1.childNodes[0].data)

   # 获取到pic
   pic = twoCellAnchor.getElementsByTagName('xdr'+':'+'pic')[0]
   blipFill = pic.getElementsByTagName('xdr'+':'+'blipFill')[0]
   blip = blipFill.getElementsByTagName('a'+':'+'blip')[0]
   if blip.hasAttribute("r"+":"+"embed"):
      pic="picture"+(blip.getAttribute("r"+":"+"embed"))[3:]

   # 获取活跃的表对象
   collect_active=collect.active
   session_active=session.active

   # 获取讲师姓名、职位、track、演讲主题、邮箱
   zh_name=collect_active.cell(row,3).value
   position=collect_active.cell(row,6).value
   track=collect_active.cell(row,9).value
   #title=collect_active.cell(row,4).value
   mail=collect_active.cell(row,10).value

   # 判断是否为中文名字
   if '\u4e00' <=zh_name <='\u9fff':
      # 将中文姓名转化成英文名
      name_list=lazy_pinyin(zh_name)
      xin=name_list[0]
      ming_list=name_list[1:]
      ming=""
      for i in ming_list:
         ming=ming+i
      en_name=ming.capitalize()+"_"+xin.capitalize()
   else:
      en_name=zh_name

   # 根据收集表得到的mail，从session表中查找到对应行,若没有对应行，返回0
   flag=False
   for cell1 in session_active['J']:
      str=cell1.value
      mail_list=str.split(",")
      for i in mail_list:
         session_mail=get_str_btw(i,'<','>')
         if session_mail==mail :
            flag=True
            session_row=cell1.row
            title=session_active.cell(cell1.row,7).value
            break
   if flag==False:
      session_row=0
      title=collect_active.cell(row,4).value       

   # 将数据写入CSV文件
   data=[
      (row,pic,zh_name,en_name,position,track,title,mail,session_row)
   ]
   with open("./mapping.csv","a+",encoding="utf-8",newline="") as f:
      csvf=csv.writer(f)
      csvf.writerows(data)




   